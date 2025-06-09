import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageDraw, ImageFont, ImageTk
from openpyxl import load_workbook
from fpdf import FPDF
from datetime import datetime
from tkinter import colorchooser
import re
import json
import threading
import sys
import platform
import colorsys



class CertificateApp:
    
    def __init__(self, root):
        self.root = root
        self.root.title("CertWizard - Professional Certificate Generator")
        # Set taskbar title
        if platform.system() == 'Windows':
            self.root.wm_attributes('-toolwindow', True)  # Remove from taskbar
            self.root.wm_attributes('-toolwindow', False)  # Add back with proper title
        self.set_icon()

        self.original_image = None
        self.display_image = None
        self.scale_x = self.scale_y = 1
        self.template_path = self.excel_path = None
        self.placeholders = {}
        self.excel_data = []
        self.fields = []
        self.field_vars = {}
        self.font_settings = {}
        self.selected_field = tk.StringVar()
        self.color_space = tk.StringVar(value="RGB")

        self.include_name = tk.BooleanVar(value=True)
        self.include_id = tk.BooleanVar(value=True)
        self.include_start = tk.BooleanVar(value=True)
        self.include_end = tk.BooleanVar(value=True)

        # Load fonts and setup UI
        self.available_fonts = self.load_available_fonts()
        
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the main UI components"""
        self.root.configure(padx=15, pady=15, bg="#f0f2f5")  # Modern light blue-gray background
        
        # Navigation bar with gradient effect
        nav_frame = tk.Frame(self.root, bg="#2c3e50", height=50)  # Dark blue-gray
        nav_frame.pack(fill="x", pady=(0, 10))
        
        # Logo/Title
        title_label = tk.Label(nav_frame, text="CertWizard", font=("Arial", 16, "bold"),
                             fg="white", bg="#2c3e50", padx=15)
        title_label.pack(side="left")
        
        # File menu with modern styling
        file_menu = tk.Menubutton(nav_frame, text="Project", bg="#2c3e50", fg="white",
                                 relief="flat", font=("Arial", 10), padx=10)
        file_menu.pack(side="left", padx=5)
        file_menu.menu = tk.Menu(file_menu, tearoff=0, bg="#ffffff", fg="#2c3e50",
                               activebackground="#3498db", activeforeground="white")
        file_menu["menu"] = file_menu.menu
        file_menu.menu.add_command(label="Save Project", command=self.save_project)
        file_menu.menu.add_command(label="Load Project", command=self.load_project)

        # Load buttons with modern styling
        for text, cmd in [("Load Template", self.load_template), ("Load Excel", self.load_excel)]:
            tk.Button(nav_frame, text=text, command=cmd, fg="white", bg="#3498db",
                     relief="flat", padx=12, pady=5, font=("Arial", 9),
                     activebackground="#2980b9").pack(side="left", padx=5)

        # Status bar with modern styling
        self.status_bar = tk.Label(self.root, text="Ready", bd=0, relief=tk.FLAT,
                                 anchor=tk.W, bg="#2c3e50", fg="white", padx=10, pady=5)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Main content area
        main_frame = tk.Frame(self.root, bg="#f0f2f5")
        main_frame.pack(fill="both", expand=True)

        # Settings panel with modern styling
        settings_frame = tk.Frame(main_frame, width=280, bg="white", padx=15, pady=15)
        settings_frame.pack(side="left", fill="y", padx=(0, 15))

        # Field toggles with modern styling
        self.toggle_frame = tk.LabelFrame(settings_frame, text="Certificate Fields",
                                        padx=10, pady=10, font=("Arial", 11, "bold"),
                                        bg="white", fg="#2c3e50")
        self.toggle_frame.pack(fill="x", pady=(0, 10))

        # Action buttons with modern styling
        action_frame = tk.Frame(settings_frame, bg="white")
        action_frame.pack(fill="x", pady=(10, 5))

        for text, cmd, color in [("Preview", self.preview_certificate, "#27ae60"),
                               ("Generate", self.generate_certificates, "#3498db")]:
            tk.Button(action_frame, text=text, command=cmd, bg=color, fg="white",
                     relief="flat", padx=15, pady=8, font=("Arial", 10, "bold"),
                     activebackground="#219a52" if color == "#27ae60" else "#2980b9"
                     ).pack(side="left" if text == "Preview" else "right", padx=2, fill="x", expand=True)

        # Progress bar with modern styling
        style = ttk.Style()
        style.configure("Modern.Horizontal.TProgressbar",
                       troughcolor="#f0f2f5",
                       background="#3498db",
                       thickness=10)
        self.progress = ttk.Progressbar(settings_frame, orient="horizontal", length=200,
                                      mode="determinate", style="Modern.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(10, 0))

        # Information text box
        info_frame = tk.LabelFrame(settings_frame, text="Generation Status",
                                 padx=10, pady=10, font=("Arial", 11, "bold"),
                                 bg="white", fg="#2c3e50")
        info_frame.pack(fill="x", pady=(10, 0))

        # Create a text widget with scrollbar
        self.info_text = tk.Text(info_frame, height=8, width=30, wrap=tk.WORD,
                               font=("Arial", 9), bg="#f8f9fa", fg="#2c3e50",
                               relief=tk.FLAT, padx=5, pady=5)
        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(info_frame, orient="vertical", command=self.info_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.info_text.configure(yscrollcommand=scrollbar.set)
        
        # Make text read-only
        self.info_text.configure(state='disabled')

        # Canvas area with modern styling
        self.canvas_frame = tk.Frame(main_frame, relief="solid", borderwidth=1,
                                   bg="white", highlightbackground="#bdc3c7")
        self.canvas_frame.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        self.canvas = tk.Canvas(self.canvas_frame, bg="white", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True, padx=2, pady=2)

    def set_icon(self):
        """Set the application icon"""
        try:
            base_path = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) 
                                      else os.path.abspath(__file__))
            
            # Try different icon formats and locations
            icon_paths = [
                os.path.join(base_path, 'certgen.ico'),
                os.path.join(base_path, 'icon.ico'),
                os.path.join(base_path, 'logo.png'),
                os.path.join(base_path, 'icon.png'),
                os.path.join(base_path, 'assets', 'certgen.ico'),
                os.path.join(base_path, 'assets', 'icon.ico'),
                os.path.join(base_path, 'assets', 'logo.png'),
                os.path.join(base_path, 'assets', 'icon.png')
            ]
            
            # For PyInstaller, also check the _MEIPASS directory
            if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                meipass_path = sys._MEIPASS
                icon_paths.extend([
                    os.path.join(meipass_path, 'certgen.ico'),
                    os.path.join(meipass_path, 'icon.ico'),
                    os.path.join(meipass_path, 'logo.png'),
                    os.path.join(meipass_path, 'icon.png'),
                    os.path.join(meipass_path, 'assets', 'certgen.ico'),
                    os.path.join(meipass_path, 'assets', 'icon.ico'),
                    os.path.join(meipass_path, 'assets', 'logo.png'),
                    os.path.join(meipass_path, 'assets', 'icon.png')
                ])
            
            icon_set = False
            for icon_path in icon_paths:
                if os.path.exists(icon_path):
                    try:
                        if platform.system() == 'Windows':
                            self.root.iconbitmap(icon_path)
                            # Also set the taskbar icon
                            self.root.wm_iconbitmap(icon_path)
                            icon_set = True
                            break
                        elif platform.system() in ['Linux', 'Darwin']:
                            img = Image.open(icon_path)
                            photo = ImageTk.PhotoImage(img)
                            self.root.iconphoto(True, photo)  # True makes it the default for all windows
                            icon_set = True
                            break
                    except Exception as e:
                        print(f"Error setting icon from {icon_path}: {e}")
                        continue
            
            if not icon_set:
                print("No suitable icon file found or could not be loaded")
        except Exception as e:
            print(f"Error setting icon: {e}")

    def update_color_space(self, event=None):
        """Update all field colors when color space changes"""
        for field in self.fields:
            if field in self.font_settings:
                current_color = self.font_settings[field]["color"].get()
                if self.color_space.get() == "RGB":
                    # Convert CMYK to RGB if needed
                    if current_color.startswith("cmyk("):
                        rgb_color = self.cmyk_to_rgb(current_color)
                        self.font_settings[field]["color"].set(rgb_color)
                else:  # CMYK
                    # Convert RGB to CMYK if needed
                    if current_color.startswith("#"):
                        cmyk_color = self.rgb_to_cmyk(current_color)
                        self.font_settings[field]["color"].set(cmyk_color)
                self.update_preview(field)

    def choose_color(self, field):
        """Open color chooser with appropriate color space"""
        current_color = self.font_settings[field]["color"].get()
        
        if self.color_space.get() == "RGB":
            # RGB color chooser
            color = colorchooser.askcolor(title=f"Choose RGB color for {field}", 
                                        initialcolor=current_color if current_color.startswith("#") else "#000000")
            if color[1]:
                self.font_settings[field]["color"].set(color[1])
        else:
            # CMYK color chooser
            if current_color.startswith("cmyk("):
                c, m, y, k = map(float, current_color[5:-1].split(","))
            else:
                c, m, y, k = 0, 0, 0, 0
            
            cmyk_window = tk.Toplevel(self.root)
            cmyk_window.title(f"Choose CMYK color for {field}")
            
            # Create CMYK sliders
            c_var = tk.DoubleVar(value=c)
            m_var = tk.DoubleVar(value=m)
            y_var = tk.DoubleVar(value=y)
            k_var = tk.DoubleVar(value=k)
            
            def update_color(*args):
                cmyk_color = f"cmyk({c_var.get():.2f},{m_var.get():.2f},{y_var.get():.2f},{k_var.get():.2f})"
                self.font_settings[field]["color"].set(cmyk_color)
                preview_label.config(bg=self.cmyk_to_rgb(cmyk_color))
            
            # Create sliders
            tk.Label(cmyk_window, text="Cyan:").grid(row=0, column=0, padx=5, pady=5)
            tk.Scale(cmyk_window, from_=0, to=1, resolution=0.01, variable=c_var, 
                    command=update_color, orient="horizontal").grid(row=0, column=1, padx=5, pady=5)
            
            tk.Label(cmyk_window, text="Magenta:").grid(row=1, column=0, padx=5, pady=5)
            tk.Scale(cmyk_window, from_=0, to=1, resolution=0.01, variable=m_var, 
                    command=update_color, orient="horizontal").grid(row=1, column=1, padx=5, pady=5)
            
            tk.Label(cmyk_window, text="Yellow:").grid(row=2, column=0, padx=5, pady=5)
            tk.Scale(cmyk_window, from_=0, to=1, resolution=0.01, variable=y_var, 
                    command=update_color, orient="horizontal").grid(row=2, column=1, padx=5, pady=5)
            
            tk.Label(cmyk_window, text="Black:").grid(row=3, column=0, padx=5, pady=5)
            tk.Scale(cmyk_window, from_=0, to=1, resolution=0.01, variable=k_var, 
                    command=update_color, orient="horizontal").grid(row=3, column=1, padx=5, pady=5)
            
            # Preview
            preview_label = tk.Label(cmyk_window, width=20, height=10)
            preview_label.grid(row=4, column=0, columnspan=2, pady=10)
            update_color()
            
            # OK button
            tk.Button(cmyk_window, text="OK", command=cmyk_window.destroy).grid(row=5, column=0, columnspan=2, pady=10)
            
            cmyk_window.transient(self.root)
            cmyk_window.grab_set()
            self.root.wait_window(cmyk_window)
        
        self.update_preview(field)

    def rgb_to_cmyk(self, rgb_color):
        """Convert RGB hex color to CMYK string"""
        if rgb_color.startswith("#"):
            r = int(rgb_color[1:3], 16) / 255
            g = int(rgb_color[3:5], 16) / 255
            b = int(rgb_color[5:7], 16) / 255
            
            k = 1 - max(r, g, b)
            if k == 1:
                c = m = y = 0
            else:
                c = (1 - r - k) / (1 - k)
                m = (1 - g - k) / (1 - k)
                y = (1 - b - k) / (1 - k)
            
            return f"cmyk({c:.2f},{m:.2f},{y:.2f},{k:.2f})"
        return rgb_color

    def cmyk_to_rgb(self, cmyk_color):
        """Convert CMYK string to RGB hex color"""
        if cmyk_color.startswith("cmyk("):
            c, m, y, k = map(float, cmyk_color[5:-1].split(","))
            
            r = 255 * (1 - c) * (1 - k)
            g = 255 * (1 - m) * (1 - k)
            b = 255 * (1 - y) * (1 - k)
            
            return f"#{int(r):02x}{int(g):02x}{int(b):02x}"
        return cmyk_color

    def hex_to_rgb(self, hex_color):
        """Convert hex color to RGB tuple, handling both RGB and CMYK"""
        if isinstance(hex_color, tk.StringVar):
            hex_color = hex_color.get()
        
        if hex_color.startswith("cmyk("):
            # Convert CMYK to RGB
            c, m, y, k = map(float, hex_color[5:-1].split(","))
            r = int(255 * (1 - c) * (1 - k))
            g = int(255 * (1 - m) * (1 - k))
            b = int(255 * (1 - y) * (1 - k))
            return (r, g, b)
        else:
            # Handle RGB hex color
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

    def load_template(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png")])
        if not file_path:
            return

        try:
            self.template_path = file_path
            self.original_image = Image.open(file_path)
            original_width, original_height = self.original_image.size

            max_width, max_height = 1000, 700
            ratio = min(max_width / original_width, max_height / original_height)
            new_size = (int(original_width * ratio), int(original_height * ratio))

            self.scale_x = original_width / new_size[0]
            self.scale_y = original_height / new_size[1]

            resized_img = self.original_image.resize(new_size)
            self.display_image = ImageTk.PhotoImage(resized_img)

            self.canvas.config(width=new_size[0], height=new_size[1])
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, image=self.display_image, anchor="nw")

            # Create placeholders for all fields
            for field in self.fields:
                self.create_placeholder(field)
        except Exception as e:
            messagebox.showerror("Error", f"Error loading template: {str(e)}")

    def create_placeholder(self, field, x=None, y=None, is_update=False):
        if field not in self.fields:
            return

        # Remove existing placeholder if any
        if field in self.placeholders:
            old_x, old_y = self.canvas.coords(self.placeholders[field])
            self.canvas.delete(self.placeholders[field])
            x = x if x is not None else old_x
            y = y if y is not None else old_y
        else:
            # For all fields, center them horizontally
            preview_img = self.render_placeholder(field)
            if preview_img:
                placeholder_width = preview_img.width()
                # Calculate centered position based on canvas width
                canvas_width = self.canvas.winfo_width()
                x = (canvas_width - placeholder_width) // 2
                y = 50  # Default vertical position

        # Create new placeholder
        preview_img = self.render_placeholder(field)
        if preview_img:
            # Store the image reference
            if not hasattr(self, '_placeholder_images'):
                self._placeholder_images = {}
            self._placeholder_images[field] = preview_img
            
            # Create text directly on canvas
            item = self.canvas.create_image(x, y, image=preview_img, anchor="center")
            self.canvas.tag_bind(item, "<Button-1>", lambda e, i=item: self.start_drag(e, i))
            self.canvas.tag_bind(item, "<B1-Motion>", lambda e, i=item: self.do_drag(e, i))
            self.placeholders[field] = item

    def start_drag(self, event, item):
        self._drag_data = {
            "item": item,
            "x": event.x,
            "y": event.y
        }

    def do_drag(self, event, item):
        dx = event.x - self._drag_data["x"]
        dy = event.y - self._drag_data["y"]
        self.canvas.move(item, dx, dy)
        self._drag_data["x"] = event.x
        self._drag_data["y"] = event.y

    def toggle_placeholder(self, field):
        if field in self.placeholders:
            if self.field_vars[field].get():
                # Recreate the placeholder with the stored image
                if hasattr(self, '_placeholder_images') and field in self._placeholder_images:
                    x, y = self.canvas.coords(self.placeholders[field])
                    self.canvas.delete(self.placeholders[field])
                    item = self.canvas.create_image(x, y, image=self._placeholder_images[field], anchor="center")
                    self.canvas.tag_bind(item, "<Button-1>", lambda e, i=item: self.start_drag(e, i))
                    self.canvas.tag_bind(item, "<B1-Motion>", lambda e, i=item: self.do_drag(e, i))
                    self.placeholders[field] = item
            else:
                self.canvas.delete(self.placeholders[field])
                del self.placeholders[field]

    def update_preview(self, field):
        if field in self.placeholders:
            # Get current coordinates before recreating placeholder
            x, y = self.canvas.coords(self.placeholders[field])
            # Create placeholder with is_update=True to prevent centering
            self.create_placeholder(field, x, y, is_update=True)
            # Update status
            self.update_status(f"Updated {field} settings")

    def get_placeholder_positions(self):
        """Get scaled coordinates for actual certificate."""
        coords = {}
        for field, item in self.placeholders.items():
            try:
                x, y = self.canvas.coords(item)
                # Scale the coordinates back to original image size
                scaled_x = x * self.scale_x
                scaled_y = y * self.scale_y
                coords[field] = (scaled_x, scaled_y)
            except Exception as e:
                print(f"Error getting coordinates for {field}: {e}")
        return coords
    
    def load_excel(self, file_path=None):
        if not file_path:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        self.excel_path = file_path

        try:
            wb = load_workbook(file_path)
            sheet = wb.active

            # Get field names from header row and normalize them
            self.fields = [str(cell.value).strip().lower() for cell in sheet[1] if cell.value]
            
            if not self.fields:
                messagebox.showerror("Error", "No fields found in Excel file!")
                return

            # Initialize field variables and font settings
            self.field_vars = {}
            self.font_settings = {}
            for field in self.fields:
                self.field_vars[field] = tk.BooleanVar(value=True)
                self.font_settings[field] = {
                    "size": tk.IntVar(value=32),
                    "color": tk.StringVar(value="#000000"),
                    "font_name": tk.StringVar(value=list(self.available_fonts.keys())[0] if self.available_fonts else "Default")
                }

            # Load data
            self.excel_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                student_data = {}
                for i, value in enumerate(row):
                    if i < len(self.fields):
                        if isinstance(value, datetime):
                            value = value.strftime("%d-%m-%Y")
                        student_data[self.fields[i]] = str(value) if value is not None else ""
                if any(student_data.values()):  # Only add if there's at least one non-empty value
                    self.excel_data.append(student_data)

            if not self.excel_data:
                messagebox.showwarning("Warning", "No data found in the Excel file!")
                return

            print(f"Loaded {len(self.excel_data)} students with fields: {', '.join(self.fields)}")
            
            # Update UI with new fields
            self.update_ui_fields()
            
            # Create placeholders for all fields
            if self.original_image:
                for field in self.fields:
                    self.create_placeholder(field)

            messagebox.showinfo("Success", f"Successfully loaded {len(self.excel_data)} records!")
        except Exception as e:
            messagebox.showerror("Error", f"Error loading Excel file: {str(e)}")

    def update_ui_fields(self):
        """Update UI elements to reflect current fields."""
        # Clear existing placeholders
        for placeholder in self.placeholders.values():
            self.canvas.delete(placeholder)
        self.placeholders.clear()
        if hasattr(self, '_placeholder_images'):
            self._placeholder_images.clear()

        # Update toggle frame
        for widget in self.toggle_frame.winfo_children():
            widget.destroy()

        # Create a scrollable frame for fields
        canvas = tk.Canvas(self.toggle_frame, height=300, bg="#ffffff", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.toggle_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#ffffff")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        for field in self.fields:
            # Create field frame with a border and padding
            field_frame = tk.Frame(scrollable_frame, relief="solid", borderwidth=1, 
                                 padx=10, pady=8, bg="#ffffff")
            field_frame.pack(fill="x", pady=2, padx=2)

            # Top row: Field name
            top_frame = tk.Frame(field_frame, bg="#ffffff")
            top_frame.pack(fill="x", pady=(0, 5))
            
            field_label = tk.Label(top_frame, text=field.title(), 
                                 font=("Arial", 10, "bold"), width=15, bg="#ffffff")
            field_label.pack(side="left", padx=2)

            # Bottom row: Font settings
            bottom_frame = tk.Frame(field_frame, bg="#ffffff")
            bottom_frame.pack(fill="x")

            # Font size
            size_frame = tk.Frame(bottom_frame, bg="#ffffff")
            size_frame.pack(side="left", padx=2)
            tk.Label(size_frame, text="Size:", bg="#ffffff").pack(side="left")
            size_spinbox = tk.Spinbox(size_frame, from_=10, to=200, 
                                    textvariable=self.font_settings[field]["size"],
                                    width=4, command=lambda f=field: self.update_preview(f))
            size_spinbox.pack(side="left")

            # Font selection
            font_frame = tk.Frame(bottom_frame, bg="#ffffff")
            font_frame.pack(side="left", padx=2)
            tk.Label(font_frame, text="Font:", bg="#ffffff").pack(side="left")
            font_option = ttk.Combobox(font_frame, values=list(self.available_fonts.keys()),
                                     textvariable=self.font_settings[field]["font_name"],
                                     width=15, state="readonly")
            font_option.pack(side="left")
            font_option.bind("<<ComboboxSelected>>", lambda e, f=field: self.update_preview(f))

            # Color button with preview
            color_frame = tk.Frame(bottom_frame, bg="#ffffff")
            color_frame.pack(side="right", padx=2)
            color_preview = tk.Label(color_frame, width=3, height=1, 
                                   bg=self.font_settings[field]["color"].get(),
                                   relief="solid", borderwidth=1)
            color_preview.pack(side="left", padx=(0, 5))
            color_btn = tk.Button(color_frame, text="Color", 
                                command=lambda f=field: self.choose_color(f),
                                relief="flat", bg="#e8e8e8", padx=8, pady=2,
                                font=("Arial", 9))
            color_btn.pack(side="left")

        # Pack the scrollbar and canvas
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Create placeholders for all fields
        if self.original_image:
            for field in self.fields:
                if self.field_vars[field].get():
                    self.create_placeholder(field)

    def ensure_font_settings_keys(self):
        for field in self.fields:
            if field not in self.font_settings:
                self.font_settings[field] = {}

            for key, default_value in [
                ("size", 30),
                ("color", "#000000"),
                ("font_name", list(self.available_fonts.keys())[0] if self.available_fonts else "Default")
            ]:
                if key not in self.font_settings[field]:
                    if key == "size":
                        self.font_settings[field][key] = tk.IntVar(value=default_value)
                    else:
                        self.font_settings[field][key] = tk.StringVar(value=default_value)

    def get_font_path(self, font_name):
        """Get the full path for a font name"""
        if font_name in self.available_fonts:
            font_path = self.available_fonts[font_name]
            if os.path.isabs(font_path):
                return font_path
            
            # Search in common locations
            search_paths = [
                "/usr/share/fonts",
                "/usr/local/share/fonts",
                "C:\\Windows\\Fonts",
                os.path.expanduser("~/.fonts")
            ]
            
            # Add fonts directory to search paths
            base_paths = [os.path.dirname(sys.executable if getattr(sys, 'frozen', False) 
                                        else os.path.abspath(__file__))]
            if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                base_paths.append(sys._MEIPASS)
            
            for base_path in base_paths:
                search_paths.insert(0, os.path.join(base_path, "fonts"))
            
            for font_dir in search_paths:
                full_path = os.path.join(font_dir, font_path)
                if os.path.exists(full_path):
                    return full_path
                    
        return "arial.ttf"

    def get_font_with_style(self, font_name, size):
        """Get the appropriate font"""
        try:
            return ImageFont.truetype(self.get_font_path(font_name), size)
        except:
            return ImageFont.load_default()

    def preview_certificate(self):
        if not self.original_image:
            messagebox.showwarning("CertWizard", "Please load a template first!")
            return

        if not self.excel_data:
            messagebox.showwarning("CertWizard", "Please load student data first!")
            return

        # Automatically select the first student
        student = self.excel_data[0]

        # Create a copy of the image to work on
        img = self.original_image.copy()
        draw = ImageDraw.Draw(img)
        placeholder_positions = self.get_placeholder_positions()

        for field in self.fields:
            if self.field_vars[field].get() and field in placeholder_positions:
                try:
                    x, y = placeholder_positions[field]
                    size = self.font_settings[field]["size"].get()
                    color = self.font_settings[field]["color"].get()
                    font_name = self.font_settings[field]["font_name"].get()
                    
                    font = self.get_font_with_style(font_name, size)
                    text = student[field]
                    text_width = draw.textlength(text, font=font)
                    
                    try:
                        bbox = font.getbbox(text)
                        text_height = bbox[3] - bbox[1]
                        y_offset = (size - text_height) // 2
                    except:
                        y_offset = 0

                    x = x - (text_width / 2)
                    y = y - (size / 2) + y_offset

                    draw.text((x, y), text, font=font, fill=self.hex_to_rgb(color))
                except Exception as e:
                    print(f"Error drawing text for {field}: {e}")

        # Create preview window as child of main window
        preview_win = tk.Toplevel(self.root)
        preview_win.title("CertWizard - Certificate Preview")
        preview_win.transient(self.root)  # Make it a child window
        preview_win.grab_set()  # Make it modal
        
        # Set window icon
        if hasattr(self.root, 'iconbitmap'):
            preview_win.iconbitmap(self.root.iconbitmap())
        
        # Calculate preview window size (80% of main window)
        main_width = self.root.winfo_width()
        main_height = self.root.winfo_height()
        preview_width = int(main_width * 0.8)
        preview_height = int(main_height * 0.8)
        
        # Calculate position to center the preview window
        x = self.root.winfo_x() + (main_width - preview_width) // 2
        y = self.root.winfo_y() + (main_height - preview_height) // 2
        
        # Set preview window size and position
        preview_win.geometry(f"{preview_width}x{preview_height}+{x}+{y}")
        
        # Create a frame for the preview
        preview_frame = tk.Frame(preview_win, bg="white")
        preview_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Resize image for display while maintaining aspect ratio
        img_width, img_height = img.size
        ratio = min(preview_width / img_width, preview_height / img_height)
        new_size = (int(img_width * ratio), int(img_height * ratio))
        preview_img = img.resize(new_size, Image.LANCZOS)
        preview_photo = ImageTk.PhotoImage(preview_img)

        # Label to display the image
        label = tk.Label(preview_frame, image=preview_photo, bg="white")
        label.image = preview_photo
        label.pack(fill="both", expand=True)

        # Add close button with modern styling
        close_btn = tk.Button(preview_win, text="Close", command=preview_win.destroy,
                            bg="#3498db", fg="white", relief="flat", padx=20, pady=5,
                            font=("Arial", 10, "bold"), activebackground="#2980b9")
        close_btn.pack(pady=10)
        
        # Bind escape key to close
        preview_win.bind('<Escape>', lambda e: preview_win.destroy())
        
        # Make preview window resizable
        preview_win.resizable(True, True)
        
        # Set minimum size
        preview_win.minsize(400, 300)
        
        # Center the window
        preview_win.update_idletasks()
        preview_win.lift()
        preview_win.focus_set()

    def generate_certificates(self):
        if not self.excel_data:
            messagebox.showwarning("CertWizard", "Please load student data first!")
            return
    
        if not self.original_image:
            messagebox.showwarning("CertWizard", "Please load a template first!")
            return

        # Ask user for color space
        color_space = messagebox.askyesno("CertWizard", 
            "Do you want to generate certificates in CMYK color space?\n\n" +
            "Yes = CMYK\nNo = RGB")
        
        self.color_space.set("CMYK" if color_space else "RGB")
    
        output_dir = filedialog.askdirectory(title="Select Output Folder")
        if not output_dir:
            return
        
        try:
            os.makedirs(f"{output_dir}/CMYK", exist_ok=True)
            os.makedirs(f"{output_dir}/RGB", exist_ok=True)
        except:
            print("Error Or folder already exists")
        
        generated_count = 0
    
        # Convert px to mm (1 px = 0.264583 mm)
        def px_to_mm(px): return px * 0.264583
    
        img_width_px, img_height_px = self.original_image.size
        pdf_width = px_to_mm(img_width_px)
        pdf_height = px_to_mm(img_height_px)
    
        # Update the progress bar and info
        def update_progress(current, total):
            progress = (current / total) * 100
            self.progress['value'] = progress
            self.update_info(f"Progress: {current}/{total} certificates ({progress:.1f}%)")
            self.root.update_idletasks()
    
        # Certificate generation function
        def generate_certificates_in_thread():
            nonlocal generated_count
            total_students = len(self.excel_data)
            placeholder_positions = self.get_placeholder_positions()
            
            # Clear and initialize info box
            self.update_info("Starting certificate generation...", clear=True)
            self.update_info(f"Total certificates to generate: {total_students}")
            self.update_info(f"Color space: {self.color_space.get()}")
            self.update_info(f"Output directory: {output_dir}")
            self.update_info("----------------------------------------")

            for i, student in enumerate(self.excel_data):
                try:
                    # Update status for current certificate
                    self.update_info(f"\nProcessing certificate {i+1}/{total_students}")
                    self.update_info(f"Student: {student[self.fields[0]]}")
                    
                    pdf = FPDF(unit="mm", format=(pdf_width, pdf_height))
                    pdf.add_page()
    
                    original_img = self.original_image.copy()
                    draw = ImageDraw.Draw(original_img)
        
                    # Add text fields with exact positioning
                    for field in self.fields:
                        if self.field_vars[field].get() and field in placeholder_positions:
                            try:
                                x, y = placeholder_positions[field]
                                size = self.font_settings[field]["size"].get()
                                color = self.font_settings[field]["color"].get()
                                font_name = self.font_settings[field]["font_name"].get()
                                
                                font = self.get_font_with_style(font_name, size)
                                text = student[field]
                                text_width = draw.textlength(text, font=font)
                                
                                try:
                                    bbox = font.getbbox(text)
                                    text_height = bbox[3] - bbox[1]
                                    y_offset = (size - text_height) // 2
                                except:
                                    y_offset = 0

                                x = x - (text_width / 2)
                                y = y - (size / 2) + y_offset

                                draw.text((x, y), text, font=font, fill=self.hex_to_rgb(color))
                            except Exception as e:
                                self.update_info(f"Warning: Error with field '{field}': {str(e)}")
    
                    temp_img_path = "temp_certificate.png"
                    original_img.save(temp_img_path)
    
                    pdf.image(temp_img_path, x=0, y=0, w=pdf_width, h=pdf_height)
    
                    temp_name = re.sub(r'[^\w\-_. ]', '', student[self.fields[0]]).strip()
                    temp_name2 = re.sub(r'[^\w\-_. ]', '', student[self.fields[1]]).strip()
                    safe_name = f"{temp_name}_{temp_name2}"

                    if self.color_space.get() == "CMYK":
                        pdf_output_path = os.path.join(f"{output_dir}/CMYK", f"{safe_name}_certificate.pdf")
                    else:
                        pdf_output_path = os.path.join(f"{output_dir}/RGB", f"{safe_name}_certificate.pdf")
                    
                    pdf.output(pdf_output_path)
                    generated_count += 1
                    self.update_info(f"Generated: {safe_name}_certificate.pdf")
    
                    if os.path.exists(temp_img_path):
                        os.remove(temp_img_path)
    
                    update_progress(i + 1, total_students)
                except Exception as e:
                    self.update_info(f"Error processing certificate {i+1}: {str(e)}")
    
            self.update_info("\n----------------------------------------")
            self.update_info(f"Generation complete! {generated_count} certificates generated.")
            self.root.after(0, notify_done)

        # Start the certificate generation in a separate thread
        threading.Thread(target=generate_certificates_in_thread, daemon=True).start()

        def notify_done():
            messagebox.showinfo("CertWizard", f"{generated_count} certificate(s) generated successfully!")

    def save_project(self):
        if not self.original_image:
            messagebox.showwarning("Warning", "No template loaded!")
            return
    
        try:
            # Ensure template_path is set
            if not hasattr(self, 'template_path') or not self.template_path:
                messagebox.showwarning("Warning", "Template path not set!")
                return

            # Get current field settings
            field_settings = {}
            for field in self.fields:
                field_settings[field] = {
                    "size": self.font_settings[field]["size"].get(),
                    "color": self.font_settings[field]["color"].get(),
                    "visible": self.field_vars[field].get(),
                    "font_name": self.font_settings[field]["font_name"].get()
                }

            # Get current positions
            positions = self.get_placeholder_positions()

            project_data = {
                "template_path": self.template_path,
                "positions": positions,
                "field_settings": field_settings,
                "excel_path": self.excel_path if hasattr(self, 'excel_path') else None,
                "color_space": self.color_space.get(),
                "version": "1.0",  # Add version for future compatibility
                "last_modified": datetime.now().isoformat(),
                "canvas_size": {
                    "width": self.canvas.winfo_width(),
                    "height": self.canvas.winfo_height()
                }
            }
    
            file_path = filedialog.asksaveasfilename(
                defaultextension=".certwiz",
                filetypes=[("CertWizard Project", "*.certwiz")],
                initialfile="certwizard_project.certwiz"
            )
            
            if file_path:
                with open(file_path, "w") as f:
                    json.dump(project_data, f, indent=4)
                messagebox.showinfo("Saved", "Project saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save project: {e}")
    
    def load_project(self):
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("CertWizard Project", "*.certwiz")],
                initialfile="certwizard_project.certwiz"
            )
            if not file_path:
                return

            with open(file_path, "r") as f:
                project_data = json.load(f)

            # Clear existing UI elements
            self.canvas.delete("all")
            self.placeholders.clear()
            if hasattr(self, '_placeholder_images'):
                self._placeholder_images.clear()

            # Load template first
            if project_data.get("template_path") and os.path.exists(project_data["template_path"]):
                self.template_path = project_data["template_path"]
                self.original_image = Image.open(self.template_path)
                original_width, original_height = self.original_image.size

                # Calculate scaling
                max_width, max_height = 1000, 700
                ratio = min(max_width / original_width, max_height / original_height)
                new_size = (int(original_width * ratio), int(original_height * ratio))

                self.scale_x = original_width / new_size[0]
                self.scale_y = original_height / new_size[1]

                # Resize and display template
                resized_img = self.original_image.resize(new_size)
                self.display_image = ImageTk.PhotoImage(resized_img)
                self.canvas.config(width=new_size[0], height=new_size[1])
                self.canvas.create_image(0, 0, image=self.display_image, anchor="nw")

            # Load Excel data if path exists
            if project_data.get("excel_path") and os.path.exists(project_data["excel_path"]):
                self.excel_path = project_data["excel_path"]
                self.load_excel(self.excel_path)

            # Load color space
            if "color_space" in project_data:
                self.color_space.set(project_data["color_space"])

            # Load field settings and positions
            if "field_settings" in project_data and "positions" in project_data:
                for field in self.fields:
                    if field in project_data["field_settings"]:
                        settings = project_data["field_settings"][field]
                        self.font_settings[field]["size"].set(settings.get("size", 30))
                        self.font_settings[field]["color"].set(settings.get("color", "#000000"))
                        self.field_vars[field].set(settings.get("visible", True))
                        self.font_settings[field]["font_name"].set(settings.get("font_name", "Arial"))

                # Create placeholders with saved positions
                positions = project_data["positions"]
                for field in self.fields:
                    if field in positions:
                        x, y = positions[field]
                        # Scale coordinates back to display size
                        scaled_x = x / self.scale_x
                        scaled_y = y / self.scale_y
                        # Create placeholder with is_update=True to prevent centering
                        self.create_placeholder(field, scaled_x, scaled_y, is_update=True)

            messagebox.showinfo("Success", "Project loaded successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load project: {e}")

    def render_placeholder(self, field):
        try:
            size = self.font_settings[field]["size"].get()
            color = self.font_settings[field]["color"].get()
            
            # Get the actual value from excel data if available
            if self.excel_data and len(self.excel_data) > 0:
                sample_value = self.excel_data[0][field]
                if not sample_value:  # If the value is empty, use the field name
                    sample_value = field
            else:
                sample_value = field
            
            font_name = self.font_settings[field]["font_name"].get()
            try:
                # Use the actual font size for rendering
                font = self.get_font_with_style(font_name, size)
            except:
                font = ImageFont.load_default()
            
            # Create a temporary image to measure text
            temp_img = Image.new("RGBA", (1, 1), (0, 0, 0, 0))
            draw = ImageDraw.Draw(temp_img)
            text_width = draw.textlength(sample_value, font=font)
            
            # Get exact text height using font metrics
            try:
                # Try to get exact font metrics if available
                ascent, descent = font.getmetrics()
                text_height = ascent + descent
            except:
                # Fallback to font size if metrics not available
                text_height = size
            
            # Create the actual placeholder image with exact text size
            img = Image.new("RGBA", (int(text_width), int(text_height)), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            
            # Calculate vertical position to center text exactly
            try:
                # Get exact text bbox
                bbox = font.getbbox(sample_value)
                y_offset = (text_height - (bbox[3] - bbox[1])) // 2
            except:
                # Fallback to simple centering if bbox not available
                y_offset = (text_height - size) // 2
            
            # Draw text with exact positioning
            draw.text((0, y_offset), sample_value, font=font, fill=self.hex_to_rgb(color))
            
            # Scale the image for display using exact scaling factors
            scaled_width = int(text_width / self.scale_x)
            scaled_height = int(text_height / self.scale_y)
            scaled_img = img.resize((scaled_width, scaled_height), Image.LANCZOS)
            
            return ImageTk.PhotoImage(scaled_img)
        except Exception as e:
            print(f"Error rendering placeholder for {field}: {e}")
            return None

    def load_available_fonts(self):
        """Load all available fonts from the fonts directory"""
        fonts = {}
        
        # Get the correct directory whether running as script or exe
        if getattr(sys, 'frozen', False):
            # Running as compiled exe
            base_path = os.path.dirname(sys.executable)
            # For PyInstaller, also check the _MEIPASS directory
            if hasattr(sys, '_MEIPASS'):
                meipass_path = sys._MEIPASS
                fonts_dir = os.path.join(meipass_path, "fonts")
                if os.path.exists(fonts_dir):
                    for file in sorted(os.listdir(fonts_dir)):
                        if file.lower().endswith(('.ttf', '.otf')):
                            font_name = os.path.splitext(file)[0]
                            fonts[font_name] = os.path.join(fonts_dir, file)
        else:
            # Running as script
            base_path = os.path.dirname(os.path.abspath(__file__))
            
        # Look for fonts in the fonts subdirectory
        fonts_dir = os.path.join(base_path, "fonts")
        
        # Create fonts directory if it doesn't exist
        if not os.path.exists(fonts_dir):
            try:
                os.makedirs(fonts_dir)
                messagebox.showinfo("Fonts Directory", 
                    "A 'fonts' directory has been created. Please add your .ttf or .otf font files there.")
            except Exception as e:
                print(f"Error creating fonts directory: {e}")
            
        # Load fonts from the fonts directory
        if os.path.exists(fonts_dir):
            for file in sorted(os.listdir(fonts_dir)):
                if file.lower().endswith(('.ttf', '.otf')):
                    font_name = os.path.splitext(file)[0]
                    fonts[font_name] = os.path.join(fonts_dir, file)
                
        if not fonts:
            messagebox.showwarning("No Fonts Found", 
                "No fonts found in the 'fonts' directory. Please add .ttf or .otf files.")
            # Add a default font
            fonts["Default"] = "arial.ttf"
        
        # Sort the fonts dictionary by keys (font names)
        sorted_fonts = dict(sorted(fonts.items(), key=lambda x: x[0].lower()))
        return sorted_fonts

    def update_status(self, message):
        """Update the status bar with a message"""
        self.status_bar.config(text=f"CertWizard: {message}")
        self.root.update_idletasks()

    def update_info(self, message, clear=False):
        """Update the information text box with a message"""
        self.info_text.configure(state='normal')
        if clear:
            self.info_text.delete(1.0, tk.END)
        self.info_text.insert(tk.END, f"{message}\n")
        self.info_text.see(tk.END)  # Scroll to the end
        self.info_text.configure(state='disabled')
        self.root.update_idletasks()



if __name__ == "__main__":
    root = tk.Tk()
    # Set the taskbar title before creating the app
    if platform.system() == 'Windows':
        root.wm_attributes('-toolwindow', True)
        root.wm_attributes('-toolwindow', False)
    app = CertificateApp(root)
    root.mainloop()
