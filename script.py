import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk, ImageFont, ImageDraw
from openpyxl import load_workbook
from fpdf import FPDF
import re
from tkinter import colorchooser
import json
import threading
import sys
import platform
import colorsys



class CertificateApp:
    
    def __init__(self, root):

        self.master = root
        self.root = root
        self.root.title("Certificate Generator")
        self.set_icon()  # Set the icon after initializing the window

        self.original_image = None
        self.display_image = None
        self.scale_x = 1
        self.scale_y = 1
        self.template_path = None  # Add template_path attribute
        self.excel_path = None  # Add excel_path attribute

        self.placeholders = {}  # Store placeholder references

        self.excel_data = []  # Will store list of student dictionaries
        self.fields = []  # List to store dynamic fields
        self.field_vars = {}  # Dictionary to store field toggle variables
        self.font_settings = {}  # Dictionary to store font settings for each field
        self.selected_field = tk.StringVar()  # For dropdown

        self.include_name = tk.BooleanVar(value=True)
        self.include_id = tk.BooleanVar(value=True)
        self.include_start = tk.BooleanVar(value=True)
        self.include_end = tk.BooleanVar(value=True)

        # Add color space selection
        self.color_space = tk.StringVar(value="RGB")  # Default to RGB

        self.setup_ui()
        
    def setup_ui(self):
        self.master.title("Certificate Generator")
        self.master.configure(padx=20, pady=20)

        # ---- Top Navigation Bar ----
        nav_frame = tk.Frame(self.master, bg="#f0f0f0", height=50)
        nav_frame.pack(fill="x", pady=(0, 10))
        
        # Right side of nav (Controls)
        controls_frame = tk.Frame(nav_frame, bg="#f0f0f0")
        controls_frame.pack(side="left", padx=10)

        # File Operations
        file_menu = tk.Menubutton(controls_frame, text="Project", bg="#ffffff", relief="flat")
        file_menu.pack(side="left", padx=5)
        file_menu.menu = tk.Menu(file_menu, tearoff=0)
        file_menu["menu"] = file_menu.menu
        file_menu.menu.add_command(label="Save Project", command=self.save_project)
        file_menu.menu.add_command(label="Load Project", command=self.load_project)

        # ---- Main Content Area ----
        main_frame = tk.Frame(self.master)
        main_frame.pack(fill="both", expand=True)

        # ---- Left Panel (Settings) ----
        settings_frame = tk.Frame(main_frame, width=250)
        settings_frame.pack(side="left", fill="y", padx=(0, 20))

        # ---- Load Data ----
        data_frame = tk.Frame(settings_frame, width=250)
        data_frame.pack(fill="x", padx=(0.10))

        load_template_btn = tk.Button(data_frame, text="Load Template", command=self.load_template, 
                                    fg="black", bg="grey", relief="flat", padx=10)
        load_template_btn.pack(side="left", padx=5)

        load_excel_btn = tk.Button(data_frame, text="Load Excel", command=self.load_excel, 
                                 fg="black", bg="grey", relief="flat", padx=10)
        load_excel_btn.pack(side="left", padx=5)

        # ---- Field Toggles ----
        self.toggle_frame = tk.LabelFrame(settings_frame, text="Fields", padx=10, pady=10)
        self.toggle_frame.pack(fill="x", pady=(0, 10))

        # ---- Font Settings ----
        self.font_frame = tk.LabelFrame(settings_frame, text="Font Settings", padx=10, pady=10)
        self.font_frame.pack(fill="x", pady=(0, 10))

        # Add color space selection
        color_space_frame = tk.LabelFrame(self.font_frame, text="Color Space")
        color_space_frame.pack(fill="x", pady=5)
        color_space_menu = ttk.Combobox(color_space_frame, textvariable=self.color_space, 
                                      values=["RGB", "CMYK", "Both"], state="readonly", width=10)
        color_space_menu.pack(side="left", padx=5, pady=5)
        color_space_menu.bind("<<ComboboxSelected>>", self.update_color_space)

        # ---- Action buttons ----
        action_frame = tk.LabelFrame(settings_frame, padx=10, pady=10)
        action_frame.pack(fill="x", pady=(10, 10))

        preview_btn = tk.Button(action_frame, text="Preview", command=self.preview_certificate, 
                              bg="#4CAF50", fg="white", relief="flat", padx=10)
        preview_btn.pack(side="left")

        generate_btn = tk.Button(action_frame, text="Generate", command=self.generate_certificates,
                               bg="#2196F3", fg="white", relief="flat", padx=10)
        generate_btn.pack(side="right")

        # ---- Center Canvas Area ----
        center_panel = tk.Frame(main_frame)
        center_panel.pack(side="left", fill="both", expand=True)

        self.canvas_frame = tk.Frame(center_panel, relief="sunken", borderwidth=2)
        self.canvas_frame.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(self.canvas_frame, bg="white")
        self.canvas.pack(fill="both", expand=True)

        # ---- Progress Bar ----
        progress_frame = tk.Frame(self.master)
        progress_frame.pack(fill="x", pady=20)

        self.progress = ttk.Progressbar(progress_frame, orient="horizontal", length=300, mode="determinate")
        self.progress.pack(side="left", padx=10, expand=True)

    def set_icon(self):
        """Set the application icon based on the operating system."""
        try:
            if getattr(sys, 'frozen', False):
                # Running as a bundle (PyInstaller)
                base_path = sys._MEIPASS
            else:
                # Running as a script
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            # Try different icon formats
            icon_paths = [
                os.path.join(base_path, 'icon.ico'),
                os.path.join(base_path, 'logo.png'),
                os.path.join(base_path, 'icon.png')
            ]
            
            icon_set = False
            for icon_path in icon_paths:
                if os.path.exists(icon_path):
                    try:
                        if platform.system() == 'Windows':
                            self.root.iconbitmap(icon_path)
                            icon_set = True
                            break
                        elif platform.system() == 'Linux':
                            img = Image.open(icon_path)
                            photo = ImageTk.PhotoImage(img)
                            self.root.tk.call('wm', 'iconphoto', self.root._w, photo)
                            icon_set = True
                            break
                        elif platform.system() == 'Darwin':  # macOS
                            self.root.iconbitmap(icon_path)
                            icon_set = True
                            break
                    except Exception as e:
                        print(f"Error setting icon from {icon_path}: {e}")
                        continue
            
            if not icon_set:
                print("No suitable icon file found or could not be loaded")
        except Exception as e:
            print(f"Error setting icon: {e}")

    def update_color_space(self, event=None):
        """Update all field colors when color space changes"""
        for field in self.fields:
            if field in self.font_settings:
                current_color = self.font_settings[field]["color"].get()
                if self.color_space.get() == "RGB":
                    # Convert CMYK to RGB if needed
                    if current_color.startswith("cmyk("):
                        rgb_color = self.cmyk_to_rgb(current_color)
                        self.font_settings[field]["color"].set(rgb_color)
                else:  # CMYK
                    # Convert RGB to CMYK if needed
                    if current_color.startswith("#"):
                        cmyk_color = self.rgb_to_cmyk(current_color)
                        self.font_settings[field]["color"].set(cmyk_color)
                self.update_preview(field)

    def choose_color(self, field):
        """Open color chooser with appropriate color space"""
        current_color = self.font_settings[field]["color"].get()
        
        if self.color_space.get() == "RGB":
            # RGB color chooser
            color = colorchooser.askcolor(title=f"Choose RGB color for {field}", 
                                        initialcolor=current_color if current_color.startswith("#") else "#000000")
            if color[1]:
                self.font_settings[field]["color"].set(color[1])
        else:
            # CMYK color chooser
            if current_color.startswith("cmyk("):
                c, m, y, k = map(float, current_color[5:-1].split(","))
            else:
                c, m, y, k = 0, 0, 0, 0
            
            cmyk_window = tk.Toplevel(self.root)
            cmyk_window.title(f"Choose CMYK color for {field}")
            
            # Create CMYK sliders
            c_var = tk.DoubleVar(value=c)
            m_var = tk.DoubleVar(value=m)
            y_var = tk.DoubleVar(value=y)
            k_var = tk.DoubleVar(value=k)
            
            def update_color(*args):
                cmyk_color = f"cmyk({c_var.get():.2f},{m_var.get():.2f},{y_var.get():.2f},{k_var.get():.2f})"
                self.font_settings[field]["color"].set(cmyk_color)
                preview_label.config(bg=self.cmyk_to_rgb(cmyk_color))
            
            # Create sliders
            tk.Label(cmyk_window, text="Cyan:").grid(row=0, column=0, padx=5, pady=5)
            tk.Scale(cmyk_window, from_=0, to=1, resolution=0.01, variable=c_var, 
                    command=update_color, orient="horizontal").grid(row=0, column=1, padx=5, pady=5)
            
            tk.Label(cmyk_window, text="Magenta:").grid(row=1, column=0, padx=5, pady=5)
            tk.Scale(cmyk_window, from_=0, to=1, resolution=0.01, variable=m_var, 
                    command=update_color, orient="horizontal").grid(row=1, column=1, padx=5, pady=5)
            
            tk.Label(cmyk_window, text="Yellow:").grid(row=2, column=0, padx=5, pady=5)
            tk.Scale(cmyk_window, from_=0, to=1, resolution=0.01, variable=y_var, 
                    command=update_color, orient="horizontal").grid(row=2, column=1, padx=5, pady=5)
            
            tk.Label(cmyk_window, text="Black:").grid(row=3, column=0, padx=5, pady=5)
            tk.Scale(cmyk_window, from_=0, to=1, resolution=0.01, variable=k_var, 
                    command=update_color, orient="horizontal").grid(row=3, column=1, padx=5, pady=5)
            
            # Preview
            preview_label = tk.Label(cmyk_window, width=20, height=10)
            preview_label.grid(row=4, column=0, columnspan=2, pady=10)
            update_color()
            
            # OK button
            tk.Button(cmyk_window, text="OK", command=cmyk_window.destroy).grid(row=5, column=0, columnspan=2, pady=10)
            
            cmyk_window.transient(self.root)
            cmyk_window.grab_set()
            self.root.wait_window(cmyk_window)
        
        self.update_preview(field)

    def rgb_to_cmyk(self, rgb_color):
        """Convert RGB hex color to CMYK string"""
        if rgb_color.startswith("#"):
            r = int(rgb_color[1:3], 16) / 255
            g = int(rgb_color[3:5], 16) / 255
            b = int(rgb_color[5:7], 16) / 255
            
            k = 1 - max(r, g, b)
            if k == 1:
                c = m = y = 0
            else:
                c = (1 - r - k) / (1 - k)
                m = (1 - g - k) / (1 - k)
                y = (1 - b - k) / (1 - k)
            
            return f"cmyk({c:.2f},{m:.2f},{y:.2f},{k:.2f})"
        return rgb_color

    def cmyk_to_rgb(self, cmyk_color):
        """Convert CMYK string to RGB hex color"""
        if cmyk_color.startswith("cmyk("):
            c, m, y, k = map(float, cmyk_color[5:-1].split(","))
            
            r = 255 * (1 - c) * (1 - k)
            g = 255 * (1 - m) * (1 - k)
            b = 255 * (1 - y) * (1 - k)
            
            return f"#{int(r):02x}{int(g):02x}{int(b):02x}"
        return cmyk_color

    def hex_to_rgb(self, hex_color):
        """Convert hex color to RGB tuple, handling both RGB and CMYK"""
        if isinstance(hex_color, tk.StringVar):
            hex_color = hex_color.get()
        
        if hex_color.startswith("cmyk("):
            # Convert CMYK to RGB
            c, m, y, k = map(float, hex_color[5:-1].split(","))
            r = int(255 * (1 - c) * (1 - k))
            g = int(255 * (1 - m) * (1 - k))
            b = int(255 * (1 - y) * (1 - k))
            return (r, g, b)
        else:
            # Handle RGB hex color
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

    def load_template(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png")])
        if not file_path:
            return

        self.template_path = file_path
        self.original_image = Image.open(file_path)
        original_width, original_height = self.original_image.size

        max_width, max_height = 1000, 700
        ratio = min(max_width / original_width, max_height / original_height)
        new_size = (int(original_width * ratio), int(original_height * ratio))

        self.scale_x = original_width / new_size[0]
        self.scale_y = original_height / new_size[1]

        resized_img = self.original_image.resize(new_size)
        self.display_image = ImageTk.PhotoImage(resized_img)

        self.canvas.config(width=new_size[0], height=new_size[1])
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, image=self.display_image, anchor="nw")

        # Create placeholders for all fields
        for field in self.fields:
            self.create_placeholder(field)

    def create_placeholder(self, field, x=None, y=None, is_update=False):
        if field not in self.fields:
            return

        # Remove existing placeholder if any
        if field in self.placeholders:
            old_x, old_y = self.canvas.coords(self.placeholders[field])
            self.canvas.delete(self.placeholders[field])
            x = x if x is not None else old_x
            y = y if y is not None else old_y
        else:
            # For the first field, center it horizontally
            if field == self.fields[0] and not is_update:
                # Get the width of the placeholder
                preview_img = self.render_placeholder(field)
                if preview_img:
                    placeholder_width = preview_img.width()
                    # Calculate centered position based on canvas width
                    canvas_width = self.canvas.winfo_width()
                    x = (canvas_width - placeholder_width) // 2
                    y = 50  # Default vertical position
            else:
                x = 50
                y = 50

        # Create new placeholder
        preview_img = self.render_placeholder(field)
        if preview_img:
            img_label = tk.Label(self.canvas, image=preview_img, bg="white")
            img_label.image = preview_img

            # For the first field, ensure it's centered horizontally
            if field == self.fields[0] and not is_update:
                # Recalculate centered position based on actual placeholder width
                placeholder_width = preview_img.width()
                canvas_width = self.canvas.winfo_width()
                x = (canvas_width - placeholder_width) // 2

            item = self.canvas.create_window(x, y, window=img_label, anchor="nw")

            def start_drag(event):
                self._drag_data = {
                    "item": item,
                    "x": self.canvas.canvasx(event.x_root - self.canvas.winfo_rootx()),
                    "y": self.canvas.canvasy(event.y_root - self.canvas.winfo_rooty())
                }

            def do_drag(event):
                new_x = self.canvas.canvasx(event.x_root - self.canvas.winfo_rootx())
                new_y = self.canvas.canvasy(event.y_root - self.canvas.winfo_rooty())
                dx = new_x - self._drag_data["x"]
                dy = new_y - self._drag_data["y"]
                self.canvas.move(self._drag_data["item"], dx, dy)
                self._drag_data["x"] = new_x
                self._drag_data["y"] = new_y

            img_label.bind("<Button-1>", start_drag)
            img_label.bind("<B1-Motion>", do_drag)

            self.placeholders[field] = item

    def toggle_placeholder(self, field):
        if field in self.placeholders:
            if self.field_vars[field].get():
                self.canvas.itemconfig(self.placeholders[field], state="normal")
            else:
                self.canvas.itemconfig(self.placeholders[field], state="hidden")

    def update_preview(self, field):
        if field in self.placeholders:
            # Get current coordinates before recreating placeholder
            x, y = self.canvas.coords(self.placeholders[field])
            # Create placeholder with is_update=True to prevent centering
            self.create_placeholder(field, x, y, is_update=True)

    def get_placeholder_positions(self):
        """Get scaled coordinates for actual certificate."""
        coords = {}
        for field, item in self.placeholders.items():
            try:
                x, y = self.canvas.coords(item)
                # Scale the coordinates back to original image size
                scaled_x = x * self.scale_x
                scaled_y = y * self.scale_y
                coords[field] = (scaled_x, scaled_y)
            except Exception as e:
                print(f"Error getting coordinates for {field}: {e}")
        return coords
    
    def load_excel(self, file_path=None):
        if not file_path:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        self.excel_path = file_path

        wb = load_workbook(file_path)
        sheet = wb.active

        # Get field names from header row and normalize them
        self.fields = [str(cell.value).strip().lower() for cell in sheet[1]]
        
        # Initialize field variables and font settings
        self.field_vars = {}
        self.font_settings = {}
        for field in self.fields:
            self.field_vars[field] = tk.BooleanVar(value=True)
            self.font_settings[field] = {
                "size": tk.IntVar(value=32),
                "color": tk.StringVar(value="#000000")
            }

        # Load data
        self.excel_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            student_data = {}
            for i, value in enumerate(row):
                if i < len(self.fields):
                    student_data[self.fields[i]] = str(value)
            self.excel_data.append(student_data)

        print(f"Loaded {len(self.excel_data)} students with fields: {', '.join(self.fields)}")
        
        # Update UI with new fields
        self.update_ui_fields()

    def update_ui_fields(self):
        """Update UI elements to reflect current fields."""
        # Clear existing placeholders
        for placeholder in self.placeholders.values():
            self.canvas.delete(placeholder)
        self.placeholders.clear()

        # Update toggle frame
        for widget in self.toggle_frame.winfo_children():
            widget.destroy()

        for field in self.fields:
            # Create field frame
            field_frame = tk.Frame(self.toggle_frame)
            field_frame.pack(fill="x", pady=2)

            # Toggle checkbox
            cb = tk.Checkbutton(field_frame, text=field.title(), variable=self.field_vars[field],
                              command=lambda f=field: self.toggle_placeholder(f))
            cb.pack(side="left", padx=5)

            # Font size
            size_frame = tk.Frame(field_frame)
            size_frame.pack(side="left", padx=5)
            tk.Label(size_frame, text="Size:").pack(side="left")
            tk.Spinbox(size_frame, from_=10, to=100, textvariable=self.font_settings[field]["size"], 
                      width=5, command=lambda f=field: self.update_preview(f)).pack(side="left", padx=2)
            
            # Color button
            color_btn = tk.Button(field_frame, text="Color", 
                                command=lambda f=field: self.choose_color(f),
                                relief="flat", bg="#f0f0f0")
            color_btn.pack(side="right")

        # Create new placeholders
        if self.original_image:
            for field in self.fields:
                self.create_placeholder(field)

    def preview_certificate(self):
        if not self.original_image:
            messagebox.showwarning("Warning", "Template not loaded!")
            return

        if not self.excel_data:
            messagebox.showwarning("Warning", "No student data loaded!")
            return

        # Automatically select the first student
        student = self.excel_data[0]

        # Create a copy of the image to work on
        img = self.original_image.copy()
        draw = ImageDraw.Draw(img)
        placeholder_positions = self.get_placeholder_positions()
        font_path = "arial.ttf"

        for field in self.fields:
            if self.field_vars[field].get() and field in placeholder_positions:
                try:
                    x, y = placeholder_positions[field]
                    size = self.font_settings[field]["size"].get()
                    color = self.font_settings[field]["color"].get()
                    try:
                        font = ImageFont.truetype(font_path, size)
                    except IOError:
                        font = ImageFont.load_default()
                    
                    # Calculate the width of the text to adjust the position
                    text_width = draw.textlength(student[field], font=font)

                    # If it's the first field, center it
                    # if field == self.fields[0]:
                    #     x = (img.width - text_width) // 2
                    x = x - 1

                    # Apply the text to the image
                    draw.text((x, y), student[field], font=font, fill=self.hex_to_rgb(color))
                except Exception as e:
                    print(f"Error drawing text for {field}: {e}")

        # Show preview in a new window
        preview_win = tk.Toplevel(self.root)
        preview_win.title("Certificate Preview")

        # Resize image for display
        preview_img = img.resize((900, int(img.height * (900 / img.width))), Image.LANCZOS)
        preview_photo = ImageTk.PhotoImage(preview_img)

        # Label to display the image in the preview window
        label = tk.Label(preview_win, image=preview_photo)
        label.image = preview_photo
        label.pack()

        preview_win.mainloop()

    def generate_certificates(self):
        if not self.excel_data:
            messagebox.showwarning("Warning", "No student data loaded!")
            return
    
        if not self.original_image:
            messagebox.showwarning("Warning", "No template image loaded!")
            return
    
        output_dir = filedialog.askdirectory(title="Select Output Folder")
        if not output_dir:
            return
        
        if self.color_space.get() == "CMYK":
            try:
                os.mkdir(f"{output_dir}/CMYK")
            except:
                print("Folder Already Exists")
        elif self.color_space.get() == "RGB":
            try:
                os.mkdir(f"{output_dir}/RGB")
            except:
                print("Folder Already Exists")
        
    
        font_path = "arial.ttf"
        generated_count = 0
    
        # Convert px to mm (1 px = 0.264583 mm)
        def px_to_mm(px): return px * 0.264583
    
        img_width_px, img_height_px = self.original_image.size
        pdf_width = px_to_mm(img_width_px)
        pdf_height = px_to_mm(img_height_px)
    
        # Update the progress bar
        def update_progressbar(current, total):
            progress = (current / total) * 100
            self.progress['value'] = progress
            self.master.update_idletasks()
    
        # Certificate generation function
        def generate_certificates_in_thread():
            nonlocal generated_count
            total_students = len(self.excel_data)
            placeholder_positions = self.get_placeholder_positions()
    
            for i, student in enumerate(self.excel_data):
                pdf = FPDF(unit="mm", format=(pdf_width, pdf_height))
                pdf.add_page()
    
                original_img = self.original_image.copy()
                draw = ImageDraw.Draw(original_img)
        
                # Add text fields
                for field in self.fields:
                    if self.field_vars[field].get() and field in placeholder_positions:
                        try:
                            x, y = placeholder_positions[field]
                            size = self.font_settings[field]["size"].get()
                            color = self.font_settings[field]["color"].get()
                            try:
                                font = ImageFont.truetype(font_path, size)
                            except IOError:
                                font = ImageFont.load_default()
                            
                            # Calculate the width of the text to adjust the position
                            text_width = draw.textlength(student[field], font=font)
    
                            # If it's the first field, center it
                            # if field == self.fields[0]:
                            #     x = (original_img.width - text_width) // 2
                            x = x - 1
    
                            # Apply the text to the image
                            draw.text((x, y), student[field], font=font, fill=self.hex_to_rgb(color))
                        except Exception as e:
                            print(f"Error drawing text for {field}: {e}")
    
                temp_img_path = "temp_certificate.png"
                original_img.save(temp_img_path)
    
                pdf.image(temp_img_path, x=0, y=0, w=pdf_width, h=pdf_height)
    
                # Use the first field as the filename
                safe_name = re.sub(r'[^\w\-_. ]', '', student[self.fields[0]]).strip()
                    
                if self.color_space.get() == "CMYK":
                    pdf_output_path = os.path.join(f"{output_dir}/CMYK", f"{safe_name}_certificate.pdf")
                else:
                    pdf_output_path = os.path.join(f"{output_dir}/RGB", f"{safe_name}_certificate.pdf")
                pdf.output(pdf_output_path)
                generated_count += 1
    
                if os.path.exists(temp_img_path):
                    os.remove(temp_img_path)
    
                # Update progress after each certificate is generated
                update_progressbar(i + 1, total_students)
    
            messagebox.showinfo("Done", f"{generated_count} certificate(s) generated successfully!")
    
        # Start the certificate generation in a separate thread
        threading.Thread(target=generate_certificates_in_thread, daemon=True).start()

    def save_project(self):
        if not self.original_image:
            messagebox.showwarning("Warning", "No template loaded!")
            return
    
        try:
            # Ensure template_path is set
            if not hasattr(self, 'template_path') or not self.template_path:
                messagebox.showwarning("Warning", "Template path not set!")
                return

            # Get current field settings
            field_settings = {}
            for field in self.fields:
                field_settings[field] = {
                    "size": self.font_settings[field]["size"].get(),
                    "color": self.font_settings[field]["color"].get(),
                    "visible": self.field_vars[field].get()
                }

            project_data = {
                "template_path": self.template_path,
                "positions": self.get_placeholder_positions(),
                "field_settings": field_settings,
                "excel_path": self.excel_path if hasattr(self, 'excel_path') else None,
                "color_space": self.color_space.get()  # Save color space
            }
    
            file_path = filedialog.asksaveasfilename(defaultextension=".certproj", filetypes=[("Certificate Project", "*.certproj")])
            if file_path:
                with open(file_path, "w") as f:
                    json.dump(project_data, f, indent=4)
                messagebox.showinfo("Saved", "Project saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save project: {e}")
    
    def load_project(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Certificate Project", "*.certproj")])
            if not file_path:
                return

            with open(file_path, "r") as f:
                project_data = json.load(f)

            # Clear existing UI elements
            self.canvas.delete("all")
            self.placeholders.clear()

            # Load template first
            if project_data.get("template_path") and os.path.exists(project_data["template_path"]):
                self.template_path = project_data["template_path"]
                self.original_image = Image.open(self.template_path)
                original_width, original_height = self.original_image.size

                # Calculate scaling
                max_width, max_height = 1000, 700
                ratio = min(max_width / original_width, max_height / original_height)
                new_size = (int(original_width * ratio), int(original_height * ratio))

                self.scale_x = original_width / new_size[0]
                self.scale_y = original_height / new_size[1]

                # Resize and display template
                resized_img = self.original_image.resize(new_size)
                self.display_image = ImageTk.PhotoImage(resized_img)
                self.canvas.config(width=new_size[0], height=new_size[1])
                self.canvas.create_image(0, 0, image=self.display_image, anchor="nw")

            # Load Excel data if path exists
            if project_data.get("excel_path") and os.path.exists(project_data["excel_path"]):
                self.excel_path = project_data["excel_path"]
                self.load_excel(self.excel_path)

            # Load color space
            if "color_space" in project_data:
                self.color_space.set(project_data["color_space"])

            # Load field settings and positions
            if "field_settings" in project_data and "positions" in project_data:
                for field in self.fields:
                    if field in project_data["field_settings"]:
                        settings = project_data["field_settings"][field]
                        self.font_settings[field]["size"].set(settings["size"])
                        self.font_settings[field]["color"].set(settings["color"])
                        self.field_vars[field].set(settings["visible"])

                # Create placeholders with saved positions
                positions = project_data["positions"]
                for field in self.fields:
                    if field in positions:
                        x, y = positions[field]
                        # Scale coordinates back to display size
                        scaled_x = x / self.scale_x
                        scaled_y = y / self.scale_y
                        # Create placeholder with is_update=True to prevent centering
                        self.create_placeholder(field, scaled_x, scaled_y, is_update=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load project: {e}")

    def render_placeholder(self, field):
        try:
            size = self.font_settings[field]["size"].get()
            color = self.font_settings[field]["color"].get()
            
            sample_value = self.excel_data[0][field] if self.excel_data else field
            
            font_path = "arial.ttf"
            try:
                scaled_font_size = max(10, int(size / self.scale_y))
                font = ImageFont.truetype(font_path, scaled_font_size)
            except IOError:
                font = ImageFont.load_default()
            
            img = Image.new("RGBA", (500, 100), (255, 255, 255, 0))
            draw = ImageDraw.Draw(img)
            draw.text((0, 0), sample_value, font=font, fill=self.hex_to_rgb(color))
            bbox = img.getbbox()
            cropped_img = img.crop(bbox)
            
            return ImageTk.PhotoImage(cropped_img)
        except Exception as e:
            print(f"Error rendering placeholder: {e}")
            return None



if __name__ == "__main__":
    root = tk.Tk()
    app = CertificateApp(root)
    root.mainloop()